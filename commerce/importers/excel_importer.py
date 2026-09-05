import csv
import io

from openpyxl import Workbook, load_workbook
from rest_framework.exceptions import ValidationError

from .validators import TEMPLATE_COLUMNS, validate_columns


def build_template_workbook():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "products"
    sheet.append(TEMPLATE_COLUMNS)
    sheet.append([
        "Exemple Produit",
        "REAL-001",
        "Electronique",
        "Accessoires",
        "Dolphin",
        "Description courte",
        "Description propre du produit",
        "199.00",
        "179.00",
        "20",
        "5",
        "Bleu",
        "M",
        "128GB",
        "250",
        "",
        "true",
        "false",
    ])
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def parse_product_file(uploaded_file):
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        text = uploaded_file.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        validate_columns(reader.fieldnames or [])
        return list(reader)
    if name.endswith(".xlsx"):
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ValidationError({"file": "Fichier vide."})
        headers = [str(value or "").strip() for value in rows[0]]
        validate_columns(headers)
        records = []
        for values in rows[1:]:
            records.append({headers[index]: value for index, value in enumerate(values)})
        return records
    raise ValidationError({"file": "Format accepte: .xlsx ou .csv."})

